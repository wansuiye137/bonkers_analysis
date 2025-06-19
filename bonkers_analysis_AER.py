import pandas as pd
import numpy as np
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 定义颜色。红色为上升，绿色为下降，黄色为新增数据
RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

csv_files = [
    'bonkers_2025-04-23.csv',
    'bonkers_2025-05-07.csv',
    'bonkers_2025-05-14.csv',
    'bonkers_2025-05-21.csv',
    'bonkers_2025-05-28.csv',
    'bonkers_2025-06-04.csv',
    'bonkers_2025-06-11.csv',
    'bonkers_2025-06-18.csv',
    'bonkers_2025-06-19.csv',
]

all_dfs = {}
unique_accounts = set()

for file in csv_files:
    df = pd.read_csv(file)
    date = pd.to_datetime(df['RunDate'].iloc[0]).date()
    #数据唯一性标识：选择Type、Account、Bank和TermMonths
    df['key'] = df[['Type','Account', 'Bank', 'TermMonths']].apply(
        lambda x: '|'.join(x.astype(str)), axis=1)
    all_dfs[date] = df

    # 收集所有唯一账户
    for key in df['key']:
        unique_accounts.add(key)

# 创建输出文件夹
output_dir = Path('savings_analysis')
output_dir.mkdir(exist_ok=True)

# 添加缺失的dates定义
dates = sorted(all_dfs.keys())

# 2. 生成AER历史表格
# 创建DataFrame存储AER历史
aer_history = pd.DataFrame(
    index=list(unique_accounts),
    columns=dates,
    dtype=float
)

# 填充AER值
for date, df in all_dfs.items():
    for key in unique_accounts:
        if key in df['key'].values:
            aer_history.loc[key, date] = df.loc[df['key'] == key, 'AER'].values[0]

# 添加账户信息列
aer_history.reset_index(inplace=True)
aer_history[['Type','Account', 'Bank', 'TermMonths']] = aer_history['index'].str.split('|', expand=True)
aer_history.drop(columns=['index'], inplace=True)

# 重新排列列顺序
cols = ['Type','Account', 'Bank', 'TermMonths'] + [date for date in dates]
aer_history = aer_history[cols]

# 保存
aer_history.to_excel(output_dir / 'AER_history.xlsx', index=False)

print(f"唯一账户数量: {len(unique_accounts)}")
print(f"生成文件:")
print(f"  - AER_history.xlsx")