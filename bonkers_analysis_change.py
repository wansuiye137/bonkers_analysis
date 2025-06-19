import pandas as pd
import numpy as np
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 定义颜色。红色为上升，绿色为下降，黄色为新增数据
RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
GREEN_FILL = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

csv_files = [
    'bonkers_2025-04-23.csv',
    'bonkers_2025-05-07.csv',
    'bonkers_2025-05-14.csv',
    'bonkers_2025-05-21.csv',
    'bonkers_2025-05-28.csv',
    'bonkers_2025-06-04.csv',
    'bonkers_2025-06-11.csv',
    'bonkers_2025-06-18.csv',
    'bonkers_2025-06-19.csv'
]

all_dfs = {}
unique_accounts = set()

for file in csv_files:
    df = pd.read_csv(file)
    date = pd.to_datetime(df['RunDate'].iloc[0]).date()
    #数据唯一性标识：选择Type、Account、Bank和TermMonths
    df['key'] = df[['Type','Account', 'Bank', 'TermMonths']].apply(
        lambda x: '|'.join(x.astype(str)), axis=1)
    all_dfs[date] = df

    # 收集所有唯一账户
    for key in df['key']:
        unique_accounts.add(key)

# 创建输出文件夹
output_dir = Path('bonkers_analysis')
output_dir.mkdir(exist_ok=True)

# 生成每周变化表格
dates = sorted(all_dfs.keys())

for i in range(1, len(dates)):
    current_date = dates[i]
    prev_date = dates[i - 1]

    current_df = all_dfs[current_date].copy()
    prev_df = all_dfs[prev_date].copy()

    # 创建新工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"Changes_bonkers_{current_date}"

    # 添加标题行
    headers = list(current_df.columns)
    headers.remove('key')
    ws.append(headers)

    # 创建前一数据的字典用于快速查找
    prev_dict = {}
    for _, row in prev_df.iterrows():
        prev_dict[row['key']] = {
            'Min': row['Min'],
            'Max': row['Max'],
            'AER': row['AER']
        }

    # 写入数据并应用样式
    for _, row in current_df.iterrows():
        row_data = [row[col] for col in headers]
        ws.append(row_data)

        current_row = ws.max_row
        key = row['key']

        # 检查是否为新增账户
        is_new = key not in prev_dict

        # 应用黄色高亮（新增账户）
        if is_new:
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = YELLOW_FILL
        else:
            # 比较Min值
            min_col = headers.index('Min') + 1
            if row['Min'] > prev_dict[key]['Min']:
                ws.cell(row=current_row, column=min_col).fill = RED_FILL
            elif row['Min'] < prev_dict[key]['Min']:
                ws.cell(row=current_row, column=min_col).fill = GREEN_FILL

            # 比较Max值
            max_col = headers.index('Max') + 1
            if row['Max'] > prev_dict[key]['Max']:
                ws.cell(row=current_row, column=max_col).fill = RED_FILL
            elif row['Max'] < prev_dict[key]['Max']:
                ws.cell(row=current_row, column=max_col).fill = GREEN_FILL

            # 比较AER值
            aer_col = headers.index('AER') + 1
            if row['AER'] > prev_dict[key]['AER']:
                ws.cell(row=current_row, column=aer_col).fill = RED_FILL
            elif row['AER'] < prev_dict[key]['AER']:
                ws.cell(row=current_row, column=aer_col).fill = GREEN_FILL

    # 保存文件
    wb.save(output_dir / f'changes_{current_date}.xlsx')

print(f"唯一账户数量: {len(unique_accounts)}")
print(f"生成文件:")
for i in range(1, len(dates)):
    print(f"  - changes_bonkers_{dates[i]}.xlsx")